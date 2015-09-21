import plotly.plotly as py

from plotly.graph_objs import Scatter, Data
from random import randint
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from .collector import Collector


def print_header():
    print('{:>3} {:>6} {:>4} {:>5} {:>5} {:>6}'.format(
        'i',
        'd_up',
        'bet',
        'level',
        'res',
        'net'))


def roll():
    n = randint(0, 1)  # player : bank ratio is not exactly 50:50!
    if n == 1:
        return 'player'
    else:
        return 'bank'


class Game():
    def __init__(self, store_data=False):
        self.gamblers = []
        self.outcome = None
        self.store_data = store_data
        if store_data:
            self.wb_created = False
            self.wb = Workbook()
            self.ws = self.wb.active

    @staticmethod
    def roll():
        n = randint(0, 1)  # player : bank ratio is not exactly 50:50!
        if n == 1:
            return 'player'
        else:
            return 'bank'

    def register(self, gamblers):  # player objects
        self.outcome = None
        self.gamblers = gamblers
        self.run()

    def run(self):
        assert self.gamblers is not None

        for gambler in self.gamblers:
            gambler.play()

        self.outcome = self.roll()
        self.notify_observers()

    def notify_observers(self):
        assert self.outcome is not None

        for gambler in self.gamblers:
            if gambler.bet_choice == self.outcome:
                amount = gambler.bet_size * 2
                gambler.update(outcome='win', reward=amount)
            else:
                gambler.update(outcome='loss')

        # if you write data to the spreadsheet before you update
        # net won't be calculated
        # maybe find a way to calc net then write data then
        # do the rest of the strategy update process? !?!!!
        self.write_xl()

    def plotz(self):
        py.sign_in(username='etsvetanov', api_key='nsyswe1pg2')
        traces = []
        num_of_rounds = len(self.gamblers[0].net_list)
        x = [i for i in range(num_of_rounds)]

        for gambler in self.gamblers:
            trace = Scatter(name=gambler.name, x=x, y=gambler.net_list)
            traces.append(trace)

        y_net_list = [trace['y'] for trace in traces]
        y_net_total = [sum(amounts) for amounts in zip(*y_net_list)]
        total_trace = Scatter(name='Total', x=x, y=y_net_total)
        traces.append(total_trace)
        data = Data(traces)
        unique_url = py.plot(data, filename='graph')

    def create_spreadsheet(self):
        self.ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        self.ws['A1'] = 'Casino'
        columns = ['play', 'level', 'index', 'bet', 'outcome', 'net']
        l = len(columns)
        for j in range(len(self.gamblers)):
            self.ws.merge_cells(start_row=1, start_column=3 + (l * j), end_row=1, end_column=7 + (l * j))
            c = self.ws.cell(row=1, column=3 + (l * j))
            c.value = self.gamblers[j].name

        row_list = [column for x in range(len(self.gamblers)) for column in columns]
        row_list.insert(0, 'outcome')
        row_list.insert(0, 'game')
        self.ws.append(row_list)
        self.wb.save('test.xlsx')

    def write_xl(self):
        # another way to write data to the spreadsheet is to move the 'writing' logic away from
        # the Game object and have each player or strategy 'submit' data to a special object
        # that collects data
        if not self.wb_created:
            self.create_spreadsheet()
            self.wb_created = True
        # if the position of gamblers in the list changes it will mess up the excel
        # consider providing some logic to the positioning

        row_list = [item for gambler
                    in self.gamblers
                    for item in
                    [gambler.bet_choice,
                     gambler.strategy.level,
                     gambler.strategy.last_index,
                     gambler.bet_size,
                     'WIN' if gambler.bet_choice == self.outcome else 'LOSS',
                     gambler.statistics['net']]]

        row_list.insert(0, self.outcome)
        row_list.insert(0, 'x')
        self.ws.append(row_list)
        for i in range(len(self.gamblers) + 1):
            cell = self.ws.cell(row=len(self.ws.rows), column=2 + (6 * i))
            cell.fill = PatternFill(patternType='solid', start_color='FFD8E4BC')
        # self.wb.save('test.xlsx')


class BaseStrategy():
    def __init__(self, coefficient=1):
        base_row = [1, 1, 1, 2, 2, 4, 6, 10, 16, 26]
        self.c = coefficient
        self.row = [i * self.c for i in base_row]
        self.i = 0
        self.double_up = False
        self.outcome = 'loss'
        self.level = 1
        self.level_target = 0  # drop to lower level after this gets to (sum(self.row) * level ) / 2
        self.last_index = 0

    def update(self, outcome, reward=None):
        self.outcome = outcome
        self.update_index()
        self.is_double()

        if reward:
            self.level_target += reward
            self.update_level()

    def update_level(self, increase=False):
        """
        you can go to a higher level if you loose the last bet in the row
        or you can go to a lower level if you win a amount equal to the
        sum of the bets in the previous level row
        """
        if increase:
            self.level += 1
            self.level_target = 0
        elif self.level_target >= ((sum(self.row) * (2 ** (self.level - 1))) / 2):
            self.level -= 1
            self.level_target = 0
        else:
            pass

        if self.level < 1:
            self.level = 1

    def update_index(self):
        self.last_index = self.i  # we are remembering the last index before calculating the new one

        if self.outcome == 'loss':
            self.i += 1
        elif self.i == 3 or self.i >= 5:
            self.i -= 3
        else:  # if self.i <= 2 or self.i == 4 AND last_outcome == 'w'
            if self.double_up:
                self.i = 0  # this is after we've played double bet and won -> we must go to 0

        if self.i >= len(self.row):  # if we loose all go to [0] 
            self.i = 0
            self.update_level(increase=True)
            # self.level += 1  # or self.level *= 2 ... linear or geometric
        assert 0 <= self.i < len(self.row)

    def is_double(self):
        # self.double_up is going to be used in THIS play
        if (self.i <= 2 or self.i == 4) and self.last_index == self.i and self.double_up is False:
            self.double_up = True
        elif self.double_up:  # if double_up was True till now
            self.double_up = False

    def get_bet_size(self):  # res - result
        level_multiplier = 2 ** (self.level - 1)

        if self.double_up:
            bet = self.row[self.i] * level_multiplier * 2  # or make double_up int and multiply by it
        else:
            bet = self.row[self.i] * level_multiplier

        assert bet > 0
        self.level_target -= bet

        return bet

    @staticmethod
    def get_bet_choice():
        choice = roll()
        return choice


class PairStrategy(BaseStrategy):
    def __init__(self, coefficient=1):
        self.lead = False
        BaseStrategy.__init__(self, coefficient)
        self.pair = None

    def set_pair(self, pair):
        self.pair = pair

    def get_bet_choice(self):
        if self.pair.strategy.lead:
            self.pair.strategy.lead = False
            if self.pair.bet_choice == 'player':
                return 'bank'
            else:
                return 'player'
        else:
            self.lead = True
            choice = roll()
            return choice


class Player():
    def __init__(self, strategy, name, cltr=None):
        self.cltr = cltr
        self.name = name
        self.strategy = strategy
        self.bet_size = None
        self.bet_choice = None
        self.statistics = {'net': 0, 'won': 0, 'lost': 0, 'largest_bet': 0}
        self.res = 'loss'
        self.net_list = []

    def print_turn(self):
        print('{:>3} {:>6} {:>4} {:>5} {:>5} {:>6}'.format(
            self.strategy.i,
            str(self.strategy.double_up),
            self.bet_size,
            self.strategy.level,
            self.res,
            self.statistics['net']))

    def submit_data(self):
        data = [self.bet_choice, self.strategy.level, self.strategy.i, self.bet_size,
                self.res, self.net]
        self.cltr.push_player_data(self.name, data)

    def update(self, outcome, reward=None):
        self.res = outcome
        if reward:
            self.statistics['won'] += 1
            assert reward is not None
            self.statistics['net'] += reward
        else:
            self.statistics['lost'] += 1

        if self.cltr:
            self.submit_data()

        self.print_turn()
        self.net_list.append(self.statistics['net'])

        self.strategy.update(outcome, reward)

    def bet(self, amount):
        if amount > self.statistics['largest_bet']:
            self.statistics['largest_bet'] = amount

        self.statistics['net'] -= amount

    def play(self):
        self.bet_size = self.strategy.get_bet_size()
        self.bet_choice = self.strategy.get_bet_choice()
        assert self.bet_size is not None
        assert self.bet_choice is not None

        self.bet(self.bet_size)  # this can actually be called without arguments


def strat_test_base():
    import datetime
    start = datetime.datetime.now().time()
    strat1 = BaseStrategy()
    strat2 = BaseStrategy()
    strat3 = BaseStrategy()
    collector = Collector()
    p1 = Player(strategy=strat1, name='P1', collector=collector)
    p2 = Player(strategy=strat2, name='P2', collector=collector)
    p3 = Player(strategy=strat3, name='P3', collector=collector)
    table = Game(store_data=True)
    players = [p1, p2, p3]

    print_header()
    for num in range(1000):
        table.register(players)
    table.plotz()
    table.wb.save('test.xlsx')
    print('times')
    print('start:', start)
    print('end:', datetime.datetime.now().time())


def strat_test_pair():
    strat1 = PairStrategy()
    strat2 = PairStrategy()
    p1 = Player(strategy=strat1, name='P1')
    p2 = Player(strategy=strat2, name='P2')
    p3 = Player(strategy=strat1, name='P3')
    p4 = Player(strategy=strat2, name='P4')
    p1.strategy.set_pair(p2)
    p2.strategy.set_pair(p1)
    p3.strategy.set_pair(p4)
    p4.strategy.set_pair(p3)
    players = [p1, p2, p3, p4]
    table = Game(store_data=True)

    for num in range(100):
        table.register(players)
    table.plotz()

strat_test_base()
# strat_test_pair()