from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from itertools import chain
from os import path


class Collector:
    def __init__(self, player_data_columns):
        self.player_data = defaultdict(list)
        self.game_data = []
        # self.player_data_columns = ['partner', 'play', 'level', 'index', 'bet', 'res', 'target', 'net']
        # self.columns = player_data_columns

    def push_player_data(self, name, data, append=True):
        """
        :param append:
        :param name: name of the player
        :param data: the data itself
        data = [partner, bet_choice, level, index, bet_size, res, net]
        """
        if append:
            self.player_data[name].append(data)
        else:
            self.player_data[name][-1] = data  # rewrites the last
            # print('debug - Collector, push_player_data()')

    def push_game_data(self, outcome):
        self.game_data.append(outcome)


class SpreadSheet:
    def __init__(self, cltr):
        self.cltr = cltr
        self.players = list(cltr.player_data.keys())
        self.wb = Workbook()
        self.ws = self.wb.active
        self.columns = cltr.player_data_columns

    def create_spreadsheet(self):
        self.init_spreadsheet()
        self.write()
        dt = datetime.now()
        stamp = '-'.join([str(dt.year), str(dt.month), str(dt.day), str(dt.hour), str(dt.minute), str(dt.second)])
        desktop_path = path.expanduser('~')
        desktop_path += "\\Desktop\\"
        self.wb.save(desktop_path + 'graph-' + stamp + '.xlsx')

    def init_spreadsheet(self):
        self.ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        self.ws['A1'] = 'Casino'
        l = len(self.columns)
        for j in range(len(self.players)):
            self.ws.merge_cells(start_row=1, start_column=3 + (l * j), end_row=1, end_column=2 + l + (l * j))
            c = self.ws.cell(row=1, column=3 + (l * j))
            c.value = self.players[j]

        row_list = [column for x in range(len(self.players)) for column in self.columns]
        row_list.insert(0, 'outcome')
        row_list.insert(0, 'game')
        self.ws.append(row_list)

    def write(self):
        data = [self.cltr.player_data[key] for key in self.players]
        rows = [chain.from_iterable(iterable) for iterable in zip(*data)]

        for outcome, row in zip(self.cltr.game_data, rows):
            self.ws.append(outcome + list(row))


        # for i in range(len(self.gamblers) + 1):
        #     cell = self.ws.cell(row=len(self.ws.rows), column=2 + (6 * i))
        #     cell.fill = PatternFill(patternType='solid', start_color='FFD8E4BC')
        # self.wb.save('test.xlsx')

